# apps/asistencia/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.utils import timezone
from django.db.models import Q
from .models import Asistencia
from .calculators import AsistenciaCalculator
from apps.talleres.models import Taller, Sesion
from apps.inscripciones.models import Inscripcion


@login_required
def registrar_asistencia(request, sesion_id):
    """
    Vista para registrar asistencia de una sesión
    Solo accesible por instructores y administradores
    """
    sesion = get_object_or_404(Sesion, id=sesion_id)
    taller = sesion.taller
    
    # Verificar permisos
    es_instructor = request.user == taller.instructor
    es_admin = request.user.rol == 'admin'
    
    if not (es_instructor or es_admin):
        messages.error(request, "No tienes permiso para registrar asistencia")
        return HttpResponseForbidden("No tienes permiso para registrar asistencia")
    
    # Obtener inscripciones confirmadas
    inscripciones = taller.inscripciones.filter(
        estado='confirmada'
    ).select_related('participante').order_by(
        'participante__first_name', 
        'participante__last_name'
    )
    
    if request.method == 'POST':
        # Procesar asistencia
        asistencias_registradas = 0
        
        for inscripcion in inscripciones:
            estado_key = f'asistencia_{inscripcion.id}'
            observaciones_key = f'observaciones_{inscripcion.id}'
            
            estado = request.POST.get(estado_key)
            observaciones = request.POST.get(observaciones_key, '')
            
            if estado:
                # Crear o actualizar asistencia
                asistencia, created = Asistencia.objects.update_or_create(
                    inscripcion=inscripcion,
                    sesion=sesion,
                    defaults={
                        'estado': estado,
                        'observaciones': observaciones
                    }
                )
                asistencias_registradas += 1
                
                # Actualizar porcentaje de asistencia
                AsistenciaCalculator.actualizar_porcentaje_inscripcion(inscripcion)
        
        messages.success(
            request, 
            f'Asistencia registrada correctamente para {asistencias_registradas} participantes'
        )
        return redirect('detalle_taller', taller_id=taller.id)
    
    # GET: Mostrar formulario
    # Obtener asistencias ya registradas
    asistencias_registradas = {}
    for inscripcion in inscripciones:
        try:
            asistencia = Asistencia.objects.get(
                inscripcion=inscripcion,
                sesion=sesion
            )
            asistencias_registradas[inscripcion.id] = asistencia
        except Asistencia.DoesNotExist:
            pass
    
    # Verificar si es posible registrar (no sesiones futuras)
    puede_registrar = sesion.fecha <= timezone.now()
    
    context = {
        'sesion': sesion,
        'taller': taller,
        'inscripciones': inscripciones,
        'asistencias_registradas': asistencias_registradas,
        'puede_registrar': puede_registrar,
        'total_inscritos': inscripciones.count(),
    }
    
    return render(request, 'asistencia/registrar_asistencia.html', context)


@login_required
def ver_asistencia_participante(request, inscripcion_id):
    """
    Vista para que un participante vea su propia asistencia
    """
    inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id)
    
    # Verificar que sea el participante o admin
    if inscripcion.participante != request.user and request.user.rol != 'admin':
        messages.error(request, "No tienes permiso para ver esta información")
        return HttpResponseForbidden("No tienes permiso para ver esta información")
    
    # Obtener asistencias ordenadas por sesión
    asistencias = inscripcion.asistencias.select_related('sesion').order_by('sesion__numero')
    
    # Calcular estadísticas
    porcentaje = AsistenciaCalculator.calcular_porcentaje(inscripcion)
    califica = AsistenciaCalculator.califica_para_certificado(inscripcion)
    
    # Contar por estado
    total_sesiones = inscripcion.taller.sesiones.count()
    presentes = asistencias.filter(estado='presente').count()
    tardes = asistencias.filter(estado='tarde').count()
    ausentes = asistencias.filter(estado='ausente').count()
    sin_registrar = total_sesiones - asistencias.count()
    
    context = {
        'inscripcion': inscripcion,
        'asistencias': asistencias,
        'porcentaje': porcentaje,
        'califica_certificado': califica,
        'total_sesiones': total_sesiones,
        'presentes': presentes,
        'tardes': tardes,
        'ausentes': ausentes,
        'sin_registrar': sin_registrar,
    }
    
    return render(request, 'asistencia/ver_asistencia.html', context)


@login_required
def estadisticas_asistencia_taller(request, taller_id):
    """
    Vista para ver estadísticas de asistencia de un taller
    Solo para instructores y administradores
    """
    taller = get_object_or_404(Taller, id=taller_id)
    
    # Verificar permisos
    es_instructor = request.user == taller.instructor
    es_admin = request.user.rol == 'admin'
    
    if not (es_instructor or es_admin):
        messages.error(request, "No tienes permiso para ver esta información")
        return HttpResponseForbidden("No tienes permiso para ver esta información")
    
    # Obtener estadísticas generales
    stats_generales = AsistenciaCalculator.obtener_estadisticas_taller(taller)
    
    # Obtener estadísticas por sesión
    sesiones = taller.sesiones.all().order_by('numero')
    stats_sesiones = []
    
    for sesion in sesiones:
        stats = AsistenciaCalculator.obtener_estadisticas_sesion(sesion)
        stats['sesion'] = sesion
        stats['sesion_pasada'] = sesion.fecha < timezone.now()
        stats_sesiones.append(stats)
    
    # Obtener lista de participantes con su porcentaje
    inscripciones = taller.inscripciones.filter(
        estado='confirmada'
    ).select_related('participante').order_by('participante__first_name')
    
    participantes_stats = []
    for inscripcion in inscripciones:
        porcentaje = AsistenciaCalculator.calcular_porcentaje(inscripcion)
        asistencias = inscripcion.asistencias.select_related('sesion').order_by('sesion__numero')
        
        participantes_stats.append({
            'inscripcion': inscripcion,
            'porcentaje': porcentaje,
            'califica': porcentaje >= 80,
            'asistencias': asistencias,
            'total_presentes': asistencias.filter(estado='presente').count(),
            'total_tardes': asistencias.filter(estado='tarde').count(),
            'total_ausentes': asistencias.filter(estado='ausente').count(),
        })
    
    context = {
        'taller': taller,
        'stats_generales': stats_generales,
        'stats_sesiones': stats_sesiones,
        'participantes_stats': participantes_stats,
        'total_sesiones': sesiones.count(),
    }
    
    return render(request, 'asistencia/estadisticas_taller.html', context)


@login_required
def api_registrar_asistencia_rapida(request):
    """
    API endpoint para registro rápido de asistencia (AJAX)
    Permite registro individual sin recargar la página
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    inscripcion_id = request.POST.get('inscripcion_id')
    sesion_id = request.POST.get('sesion_id')
    estado = request.POST.get('estado')
    observaciones = request.POST.get('observaciones', '')
    
    # Validar parámetros
    if not all([inscripcion_id, sesion_id, estado]):
        return JsonResponse({'error': 'Faltan parámetros requeridos'}, status=400)
    
    if estado not in ['presente', 'ausente', 'tarde']:
        return JsonResponse({'error': 'Estado inválido'}, status=400)
    
    try:
        inscripcion = Inscripcion.objects.get(id=inscripcion_id)
        sesion = Sesion.objects.get(id=sesion_id)
        
        # Verificar permisos
        es_instructor = request.user == sesion.taller.instructor
        es_admin = request.user.rol == 'admin'
        
        if not (es_instructor or es_admin):
            return JsonResponse({'error': 'Sin permisos'}, status=403)
        
        # Registrar asistencia
        asistencia, created = Asistencia.objects.update_or_create(
            inscripcion=inscripcion,
            sesion=sesion,
            defaults={
                'estado': estado,
                'observaciones': observaciones
            }
        )
        
        # Actualizar porcentaje
        porcentaje = AsistenciaCalculator.actualizar_porcentaje_inscripcion(inscripcion)
        
        return JsonResponse({
            'success': True,
            'created': created,
            'asistencia_id': asistencia.id,
            'porcentaje': float(porcentaje),
            'califica': porcentaje >= 80,
            'mensaje': 'Asistencia registrada' if created else 'Asistencia actualizada'
        })
        
    except Inscripcion.DoesNotExist:
        return JsonResponse({'error': 'Inscripción no encontrada'}, status=404)
    except Sesion.DoesNotExist:
        return JsonResponse({'error': 'Sesión no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def editar_asistencia(request, asistencia_id):
    """
    Vista para editar una asistencia ya registrada
    """
    asistencia = get_object_or_404(Asistencia, id=asistencia_id)
    
    # Verificar permisos
    es_instructor = request.user == asistencia.inscripcion.taller.instructor
    es_admin = request.user.rol == 'admin'
    
    if not (es_instructor or es_admin):
        messages.error(request, "No tienes permiso para editar esta asistencia")
        return HttpResponseForbidden("No tienes permiso")
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        observaciones = request.POST.get('observaciones', '')
        
        if nuevo_estado in ['presente', 'ausente', 'tarde']:
            asistencia.estado = nuevo_estado
            asistencia.observaciones = observaciones
            asistencia.save()
            
            # Actualizar porcentaje
            AsistenciaCalculator.actualizar_porcentaje_inscripcion(asistencia.inscripcion)
            
            messages.success(request, 'Asistencia actualizada correctamente')
        else:
            messages.error(request, 'Estado inválido')
        
        return redirect('registrar_asistencia', sesion_id=asistencia.sesion.id)
    
    context = {
        'asistencia': asistencia,
    }
    
    return render(request, 'asistencia/editar_asistencia.html', context)


@login_required
def exportar_asistencia_excel(request, taller_id):
    """
    Exporta reporte de asistencia a Excel
    """
    taller = get_object_or_404(Taller, id=taller_id)
    
    # Verificar permisos
    es_instructor = request.user == taller.instructor
    es_admin = request.user.rol == 'admin'
    
    if not (es_instructor or es_admin):
        return HttpResponseForbidden("No tienes permiso")
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de Asistencia - {taller.titulo}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    headers = ['Participante', 'Email', 'Programa', 'Asistencia (%)', 'Califica']
    ws.append([])
    ws.append(headers)
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    inscripciones = taller.inscripciones.filter(estado='confirmada').select_related('participante')
    
    for inscripcion in inscripciones:
        porcentaje = AsistenciaCalculator.calcular_porcentaje(inscripcion)
        califica = "Sí" if porcentaje >= 80 else "No"
        
        ws.append([
            inscripcion.participante.get_full_name(),
            inscripcion.participante.email,
            inscripcion.participante.programa_academico,
            f"{porcentaje}%",
            califica
        ])
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=asistencia_taller_{taller.id}.xlsx'
    
    wb.save(response)
    return response